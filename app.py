import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Crear libro de trabajo
wb = openpyxl.Workbook()

# Definir paleta de colores y estilos profesionales
HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")  # Slate oscuro
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
ZEBRA_FILL = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

title_font = Font(name="Calibri", size=16, bold=True, color="1F2937")
subtitle_font = Font(name="Calibri", size=11, italic=True, color="4B5563")

# 1. Pestaña de referencia con TODOS los pilotos de la parrilla
ws_ref = wb.active
ws_ref.title = "Pilotos"
ws_ref.sheet_view.showGridLines = True

drivers = [
    "Max Verstappen", "Sergio Pérez", "Lewis Hamilton", "George Russell",
    "Charles Leclerc", "Carlos Sainz", "Lando Norris", "Oscar Piastri",
    "Fernando Alonso", "Lance Stroll", "Pierre Gasly", "Esteban Ocon",
    "Alex Albon", "Yuki Tsunoda", "Daniel Ricciardo", "Valtteri Bottas",
    "Zhou Guanyu", "Kevin Magnussen", "Nico Hülkenberg", "Oliver Bearman", "Franco Colapinto"
]

ws_ref["A1"] = "Piloto"
ws_ref["A1"].font = HEADER_FONT
ws_ref["A1"].fill = HEADER_FILL
ws_ref["A1"].alignment = Alignment(horizontal="center")

for idx, driver in enumerate(drivers, start=2):
    cell = ws_ref.cell(row=idx, column=1, value=driver)
    cell.font = Font(name="Calibri", size=11)
    cell.border = thin_border
    if idx % 2 == 0:
        cell.fill = ZEBRA_FILL

ws_ref.column_dimensions['A'].width = 25

# Crear la regla de validación de datos apuntando a la hoja "Pilotos"
dv = DataValidation(type="list", formula1="=Pilotos!$A$2:$A$22", allow_blank=True)
dv.error = 'El valor introducido no está en la lista'
dv.errorTitle = 'Entrada no válida'
dv.prompt = 'Selecciona un piloto de la lista'
dv.promptTitle = 'Selección de Piloto'

# 2. Configurar las pestañas del libro
tabs_config = [
    ("Dashboard", "Panel General de Rendimiento"),
    ("Resumen", "Resumen por Gran Premio"),
    ("Rendimiento", "Métricas y Tiempos por Vuelta"),
    ("Estrategia", "Estrategia de Neumáticos y Paradas"),
    ("Comparativa", "Comparativa Cara a Cara de Pilotos")
]

for tab_name, title_text in tabs_config:
    ws = wb.create_sheet(title=tab_name)
    ws.sheet_view.showGridLines = True
    
    # Agregar la validación de datos a la hoja
    ws.add_data_validation(dv)
    
    # Bloque de Título
    ws["A1"] = title_text
    ws["A1"].font = title_font
    ws["A2"] = "Selecciona el piloto en la columna correspondiente."
    ws["A2"].font = subtitle_font
    
    # Definir cabeceras según la pestaña
    if tab_name == "Dashboard":
        headers = ["Posición", "Piloto Principal", "Equipo", "Puntos", "Velocidad Media (km/h)", "Estado"]
    elif tab_name == "Resumen":
        headers = ["Ronda", "Gran Premio", "Piloto Seleccionado", "Posición Salida", "Posición Llegada", "Puntos Obtenidos"]
    elif tab_name == "Rendimiento":
        headers = ["Sesión", "Piloto", "Mejor Vuelta", "Sector 1 (s)", "Sector 2 (s)", "Sector 3 (s)"]
    elif tab_name == "Estrategia":
        headers = ["Stint", "Piloto", "Compuesto", "Vueltas en Stint", "Degradación Est. (%)", "Pit Stop (s)"]
    else:  # Comparativa
        headers = ["Métrica", "Piloto A", "Piloto B", "Diferencia", "Ventaja"]

    header_row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    ws.row_dimensions[header_row].height = 24

    # Generar filas de ejemplo y aplicar el desplegable en las columnas de pilotos
    for r_idx in range(5, 15):
        ws.row_dimensions[r_idx].height = 20
        is_even = (r_idx % 2 == 0)
        row_fill = ZEBRA_FILL if is_even else WHITE_FILL
        
        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = Font(name="Calibri", size=11)
            cell.fill = row_fill
            cell.border = thin_border
            
            # Identificar qué columna corresponde al piloto para aplicarle la validación (dv)
            if tab_name == "Comparativa" and c_idx in [2, 3]:
                cell.value = drivers[(r_idx - 5 + c_idx) % len(drivers)]
                dv.add(cell)
            elif tab_name != "Comparativa" and c_idx == 2:
                cell.value = drivers[(r_idx - 5) % len(drivers)]
                dv.add(cell)
            elif c_idx == 1:
                cell.value = r_idx - 4
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.value = "-"
                cell.alignment = Alignment(horizontal="center")

    # Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 18)

    ws.freeze_panes = "A5"

# Guardar el archivo final
file_path = "Formula1_Parrilla_Completa.xlsx"
wb.save(file_path)
print(f"Archivo guardado exitosamente como: {file_path}")
